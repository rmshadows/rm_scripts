import os
from openpyxl import load_workbook
from docx import Document
import io

def sanitize_filename(name):
    # 清理文件名，去掉特殊字符
    return "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in name)

def get_image_ext(data):
    # 通过文件头判断图片类型（jpg, png, gif, bmp等）
    if data[:2] == b'\xff\xd8':  # JPEG 图片
        return "jpg"
    elif data[:8] == b'\x89PNG\r\n\x1a\n':  # PNG 图片
        return "png"
    elif data[:6] == b'GIF87a' or data[:6] == b'GIF89a':  # GIF 图片
        return "gif"
    elif data[:2] == b'BM':  # BMP 图片
        return "bmp"
    else:
        return "png"  # 默认使用 png

def extract_images_from_excel(file_path, output_dir):
    wb = load_workbook(file_path)
    os.makedirs(output_dir, exist_ok=True)
    print(f"正在处理 Excel 文件: {file_path}")
    for sheet in wb.worksheets:
        for image in sheet._images:
            anchor = image.anchor
            row = anchor._from.row + 1
            col = anchor._from.col + 1
            cell_value = sheet.cell(row=row, column=col).value or f"{sheet.title}_{row}_{col}"
            image_data = image._data()
            img_ext = get_image_ext(image_data)
            safe_name = sanitize_filename(str(cell_value))
            img_path = os.path.join(output_dir, f"{safe_name}.{img_ext}")
            with open(img_path, "wb") as f:
                f.write(image_data)
            print(f"导出图片: {img_path}")
    print("Excel 图片提取完成")

def extract_images_from_docx(file_path, output_dir):
    doc = Document(file_path)
    os.makedirs(output_dir, exist_ok=True)
    print(f"正在处理 Word 文件: {file_path}")
    for i, rel in enumerate(doc.part.rels.values()):
        if "image" in rel.target_ref:
            img_data = rel.target_part.blob
            img_ext = get_image_ext(img_data)
            img_name = f"image_{i+1}"
            img_path = os.path.join(output_dir, f"{sanitize_filename(img_name)}.{img_ext}")
            with open(img_path, "wb") as f:
                f.write(img_data)
            print(f"导出图片: {img_path}")
    print("Word 图片提取完成")

if __name__ == "__main__":
    import sys
    file_path = sys.argv[1] if len(sys.argv) > 1 else "1.xlsx"
    output_dir = "images_output"
    
    if file_path.endswith(".xlsx"):
        extract_images_from_excel(file_path, output_dir)
    elif file_path.endswith(".docx"):
        extract_images_from_docx(file_path, output_dir)
    else:
        print("不支持的文件格式，仅支持 .xlsx 和 .docx 文件。")

