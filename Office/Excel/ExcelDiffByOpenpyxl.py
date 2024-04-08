from openpyxl import load_workbook
from openpyxl.workbook import Workbook

#### 对比Excel文件

# 文件对比
file1 = "1.xlsx"
file2 = "2.xlsx"
# 显示单元格名称
SHOW_CELL_NAME = True
# 是否保存差异
SAVE_DIFF = None
# SAVE_DIFF = "Diff.xlsx"


def get_cell_name(row, column):
    """
    将行号和列号转换为单元格名称
    Args:
        row: 行号
        column: 列号

    Returns:
        单元格名称，例如 "A1"
    """
    # 列号转换为字母
    column_name = chr(ord('A') + column - 1)
    # 返回单元格名称
    return f"{column_name}{row}"


def compare_excel_files(file1, file2):
    # 加载 Excel 文件
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    # 获取工作表
    ws1 = wb1.active
    ws2 = wb2.active

    # 比较单元格值
    differences = []
    for row in range(1, ws1.max_row + 1):
        for col in range(1, ws1.max_column + 1):
            cell1 = ws1.cell(row=row, column=col)
            cell2 = ws2.cell(row=row, column=col)
            if cell1.value != cell2.value:
                differences.append((row, col, cell1.value, cell2.value))

    # 打印差异
    if differences:
        print("两个 Excel 文件不同，存在差异。")
        print("差异信息：")
        if SHOW_CELL_NAME:
            # 显示单元格名称
            for diff in differences:
                cell_name = get_cell_name(diff[0], diff[1])
                print(f"{cell_name}: {diff[2]} != {diff[3]}")
        else:
            for diff in differences:
                print(f"第{diff[0]}行，第{diff[1]}列：{diff[2]} != {diff[3]}")
    else:
        print("两个 Excel 文件相同，没有差异。")

    # 可选：将差异信息保存到新的 Excel 文件中
    if differences and (SAVE_DIFF is not None):
        wb_diff = Workbook()
        ws_diff = wb_diff.active
        for diff in differences:
            ws_diff.cell(row=diff[0], column=diff[1], value=f"{diff[2]} != {diff[3]}")
        wb_diff.save(SAVE_DIFF)
        print("已保存差异到："+SAVE_DIFF)

    # 可选：将差异信息保存到新的 Excel 文件中
    # if differences:
    #     wb_diff = Workbook()
    #     ws_diff = wb_diff.active
    #     for diff in differences:
    #         cell_name = get_cell_name(diff[0], diff[1])
    #         ws_diff[cell_name] = f"{diff[2]} != {diff[3]}"
    #     wb_diff.save("差异2.xlsx")


if __name__ == '__main__':
    compare_excel_files(file1, file2)

