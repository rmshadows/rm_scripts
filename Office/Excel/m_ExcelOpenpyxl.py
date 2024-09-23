#!/usr/bin/python3
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WorkbookClass  # 为了兼容 openpyxl 3.x 以及 2.x 的命名差异
from openpyxl.utils import range_boundaries

import m_Array
import m_System

from openpyxl import load_workbook, Workbook
from copy import copy

"""
文档：https://openpyxl.readthedocs.io/en/stable/tutorial.html#accessing-one-cell

#### 访问单元格
访问单元格：ws['A4'] 或者 ws.cell(row=4, column=2, value=10)
多个单元格：cell_range = ws['A1':'C2']
>>> colC = ws['C']
>>> col_range = ws['C:D']
>>> row10 = ws[10]
>>> row_range = ws[5:10]

### 所有行
Worksheet.rows

"""


def createNewWorkbook(firstSheetName="Sheet1"):
    """
    新建workbook但不是新建一个实际的文件
    获取第一张自动生成的表wb.active
    Args:
        firstSheetName: 第一Sheet表的名称
    Returns:

    """
    if firstSheetName is None:
        return Workbook()
    else:
        wb = Workbook()
        wb.active.title = firstSheetName
        return wb


def createSheet(wb, sheetName, position=None):
    """
    新增表格
    position = None # insert at the end (default)
     0) # insert at first position
     -1) # insert at the penultimate position
    默认名称：Sheet, Sheet1, Sheet2
    访问：wb["New Title"]
    Args:
        wb:
        sheetName: sheet名
        position: 位置

    Returns:
        返回新增的worksheet
    """
    if position is None:
        return wb.create_sheet(sheetName)
    else:
        return wb.create_sheet(sheetName, position)


def getSheetNames(wb):
    """
    返回Sheets名字
    Args:
        wb:workbook

    Returns:
        str
    """
    return wb.sheetnames


def getSheetByIndex(wb, index):
    """
    返回Sheet
    Args:
        wb:workbook
        index:索引
    Returns:
        str
    """
    return wb[getSheetNames(wb)[index]]


def copyWorksheet(wb, ws):
    """
    复制工作表
    Args:
        wb: 复制到某workbook
        ws: 要被复制的worksheet

    Returns:

    """
    return wb.copy_worksheet(ws)

def isNumber(var):
    return isinstance(var, (int, float))
def isString(var):
    return isinstance(var, str)


def getWorksheet(wb, indexOrName):
    """
    获取worksheet
    Args:
        wb:
        indexOrName:

    Returns:

    """
    if isNumber(indexOrName):
        # 如果指定worksheet
        return getSheetByIndex(wb, indexOrName)
    elif isString(indexOrName):
        # 如果指定名称
        return wb[indexOrName]
    else:
        return getSheetByIndex(wb, 0)


def readRows(ws, min_row, max_row, min_col, max_col, valueOnly=True):
    """
    读行
    Args:
        ws: worksheet
        min_row: 开始行
        max_row: 结束行
        min_col: 开始列
        max_col: 结束列
        valueOnly: 是否仅返回值
    Returns:
        <Cell Sheet1.A1>
        <Cell Sheet1.A2>
    """
    cs = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=valueOnly):
        for cell in row:
            print(cell)
            cs.append(cell)
    return cs


def readColumns(ws, min_col, max_col, min_row, max_row, valueOnly=True):
    """
        读列
        Args:
            ws: worksheet
            min_col: 开始列
            max_col: 结束列
            min_row: 开始行
            max_row: 结束行
            valueOnly: 是否仅返回值
        Returns:
            <Cell Sheet1.A1>
            <Cell Sheet1.A2>
        """
    cs = []
    for col in ws.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=valueOnly):
        for cell in col:
            print(cell)
            cs.append(cell)
    return cs


def getCellRange(ws, cell1, cell2):
    """
    返回一个范围内的单元格
    colC = ws['C']
    col_range = ws['C:D']
    row10 = ws[10]
    row_range = ws[5:10]
    Args:
        ws: worksheet
        ws['A1':'C2']
        cell1: 起始
        cell2: 末尾

    Returns:

    """
    return ws[cell1:cell2]


def getRowOrColumns(ws, name):
    """
    返回行或者列
    colC = ws['C']
    col_range = ws['C:D']
    row10 = ws[10]
    row_range = ws[5:10]
    Args:
        ws: worksheet
        ws['A1':'C2']

    Returns:

    """
    return ws[name]


def isnumeric(value):
    """
    检查给定参数类型
    Args:
        value:

    Returns:

    """
    if isinstance(value, int) or isinstance(value, float):
        return True
    else:
        return False


def justReadOneRow(ws, row, deep=60, valueOnly=True):
    """
    读1行,50列
    Args:
        ws: worksheet
        row: 行
        valueOnly: 是否仅返回值
    Returns:
        <Cell Sheet1.A1>
        <Cell Sheet1.A2>
    """
    cs = []
    for row in ws.iter_rows(min_row=row, max_row=row, min_col=0, max_col=deep, values_only=valueOnly):
        for cell in row:
            print(cell)
            cs.append(cell)
    return cs


def excel_column_to_number(column_name):
    """
    列名转换为对应的列索引
    Args:
        column_name:

    Returns:

    """
    result = 0
    for char in column_name:
        result = result * 26 + ord(char) - ord('A') + 1
    return result


def justReadOneColumn(ws, col, deep=50, valueOnly=True):
    """
        读列
        Args:
            ws: worksheet
            col: 列
            deep: 深度
            valueOnly: 是否仅返回值
        Returns:
            <Cell Sheet1.A1>
            <Cell Sheet1.A2>
        """
    cs = []
    if not isnumeric(col):
        col = excel_column_to_number(col)
    for col in ws.iter_cols(min_row=0, max_row=deep, min_col=col, max_col=col, values_only=valueOnly):
        for cell in col:
            print(cell)
            cs.append(cell)
    return cs


def loadExcel(filename):
    """
    加载现有的Excel表格
    Args:
        filename:Excel文件名

    Returns:
        wb
    """
    return load_workbook(filename=filename)


def replaceOneCellValue(wb, ws, cell, checkValue, replacement, whenEqual=True):
    """
    替换单元格值
    Args:
        wb: workbook
        ws: 工作表名字（字符串）或者索引（数字）
        cell: 单元格名称
        checkValue: 检查的值
        replacement: 替换值
        whenEqual: 什么时候替换，True表示是某个值的时候替换

    Returns:
        修改 返回 wb ，没有返回None
    """
    modify = False
    if isnumeric(ws):
        wss = getSheetByIndex(wb, ws)
    else:
        wss = wb[ws]
    ck = str(wss[cell].value)
    if wss[cell].value is None:
        ck = ""
    # 不论怎样直接替换
    if checkValue is None:
        wss[cell] = replacement
        modify = True
    else:
        if whenEqual:
            if ck == str(checkValue):
                wss[cell] = replacement
                modify = True
        else:
            if ck != str(checkValue):
                wss[cell] = replacement
                modify = True
    if modify:
        return wb
    else:
        return None


def merge_xlsx_files(src_folder, output_file):
    """
    创建一个新的工作簿作为合并后的文件
    (仅合并到一张worksheet！)
    Args:
        src_folder:
        output_file:

    Returns:

    """
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    for filename in os.listdir(src_folder):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(src_folder, filename)
            # 打开每个源文件
            wb = load_workbook(filepath)
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                # 复制每个工作表的内容到合并后的工作簿中
                for row in ws.iter_rows():
                    merged_ws.append([cell.value for cell in row])
    # 保存合并后的工作簿到输出文件
    merged_wb.save(output_file)


def delete_worksheets_except_index(input_file, index_to_keep, saveAs=None):
    """
    删除除了指定索引的worksheet以外的其他表格
    Args:
        input_file:文件
        index_to_keep:保留的index
        saveAs: 保存路径

    Returns:

    """
    # 加载 Excel 文件
    wb = load_workbook(input_file)
    # 获取要删除的工作表列表
    sheets_to_delete = [sheet for i, sheet in enumerate(wb.worksheets) if i != index_to_keep]
    # 删除工作表
    for sheet in sheets_to_delete:
        wb.remove(sheet)
    if saveAs is None:
        # 保存修改后的 Excel 文件
        wb.save(input_file)
    else:
        wb.save(saveAs)


def worksheetSplitInto(inputFile, outputDir=None, prefix=None):
    """
    将一个Excel中的多个worksheet拆成单个文件
    Args:
        inputFile: 输入文件
        outputDir: 输出文件夹
        prefix: 前缀

    Returns:

    """
    wb = loadExcel(inputFile)
    # 获取工作表名
    wss = getSheetNames(wb)
    # 没有文件夹就新建
    if outputDir is None:
        # 如果没给就以Excel的名字新建文件夹
        t1, t2, t3 = m_System.splitFilePath(inputFile)
        outputDir = os.path.join(t1, t2)
        print("输出文件夹: {}".format(outputDir))
    if not os.path.exists(outputDir):
        os.mkdir(outputDir)
    for i in range(0, len(wss)):
        cpf = os.path.join(outputDir, wss[i])
        # 首先复制文件
        nfn = m_System.editFilename(inputFile, cpf, prefix=prefix)
        m_System.copyFD(inputFile, nfn)
        delete_worksheets_except_index(nfn, i)


def appendData(ws, list):
    """
    添加一行数据
    we’ll enter this data onto the worksheet.
    As this is a list of lists, we can simply use the Worksheet.append() function.
    Args:
        ws: worksheet
        list: 列表

    Returns:
        none
    """
    ws.append(list)


def cellBold(ws, cellID):
    """
    字体加粗
    Args:
        ws: 工作表
        cellID: ID

    Returns:

    """
    ft = Font(bold=True)
    ws[cellID].font = ft


def barChart(title="标题", x="X轴", y="Y轴", type="col"):
    """
    返回柱状图（空的）
    Args:
        title:
        x:
        y:
        type:

    Returns:

    """
    from openpyxl.chart import BarChart
    chart = BarChart()
    chart.type = type
    chart.title = title
    chart.y_axis.title = y
    chart.x_axis.title = x
    chart.legend = None
    return chart


def reference(worksheet=None, range_string=None, min_col=None,min_row=None,max_col=None,max_row=None):
    """
    返回引用区间
    Args:
        worksheet:
        range_string:
        min_col:
        min_row:
        max_col:
        max_row:

    Returns:

    """
    from openpyxl.chart import Reference
    return Reference(worksheet, min_col, min_row, max_row, max_col, range_string)


def barChartAppendData(chart, data, categories):
    """
    为柱状统计图添加数据
    Args:
        chart:
        data: # data = Reference(ws, min_col=3, min_row=2, max_row=4, max_col=3)
        categories: # categories = Reference(ws, min_col=1, min_row=2, max_row=4, max_col=1)
    Returns:

    """
    chart.add_data(data)
    chart.set_categories(categories)
    return chart


def deleteWorksheet(wb, worksheetNameOrIndex):
    """
    删除指定的worksheet
    Args:
        wb:
        worksheetNameOrIndex: 数字索引or字符串名称

    Returns:

    """
    # 检查给定的 wb 是否是 Workbook 类型
    if not isinstance(wb, Workbook):
        raise ValueError("The 'wb' argument must be a Workbook object.")
    # 如果给定的是字符串，表示按照工作表名称删除
    if isinstance(worksheetNameOrIndex, str):
        if worksheetNameOrIndex in wb.sheetnames:
            sheet = wb[worksheetNameOrIndex]
            wb.remove(sheet)
        else:
            print(f"Worksheet '{worksheetNameOrIndex}' not found. No action taken.")
    # 如果给定的是整数，表示按照工作表索引删除
    elif isinstance(worksheetNameOrIndex, int):
        if 0 <= worksheetNameOrIndex < len(wb.sheetnames):
            wb.remove(wb.worksheets[worksheetNameOrIndex])
        else:
            print(f"Index '{worksheetNameOrIndex}' is out of range. No action taken.")
    # 其他类型不做处理
    else:
        print("Invalid input type. Only str (worksheet name) or int (worksheet index) allowed. No action taken.")
    return wb


def is_openpyxl_workbook(obj):
    """
    判断是不是workbook对象
    Args:
        obj:

    Returns:

    """
    return isinstance(obj, Workbook) or isinstance(obj, WorkbookClass)


def copyWorksheetIntoBase(wbSrcf, wsSrc, wbDstf, savePath="", copyTitle=None):
    """
    复制某Excel建议用另一个
    Args:
        wbSrcf:文件1
        wsSrc:表
        wbDstf: 复制到文件2
        savePath: 保存为
        copyTitle: 复制的标题

    Returns:

    """
    if is_openpyxl_workbook(wbSrcf):
        wbSrc = wbSrcf
    else:
        wbSrc = load_workbook(wbSrcf)
    if is_openpyxl_workbook(wbSrcf):
        wbDst = wbDstf
    else:
        wbDst = load_workbook(wbDstf)
    # 获取源工作表
    source_sheet = getWorksheet(wbSrc, wsSrc)
    # 获取工作表名字，并判断是否要修改
    wsDstName = copyTitle
    if wsDstName is None:
        wsDstName = m_Array.modify_string_if_duplicate(source_sheet.title, getSheetNames(wbDst))
        if wsDstName is None:
            wsDstName = source_sheet.title
    # 新建一张表
    target_sheet = wbDst.create_sheet(title=wsDstName)
    print(f"Copy {wbSrcf} - {source_sheet.title} -> {wbDstf} - {wsDstName}")
    # 复制数据和格式
    for row in source_sheet.iter_rows():
        for cell in row:
            target_sheet[cell.coordinate].value = cell.value
            if cell.has_style:
                target_sheet[cell.coordinate]._style = cell._style
            if cell.comment:
                target_sheet[cell.coordinate].comment = cell.comment
    # 复制合并单元格
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))
    # 调整列宽
    for col in source_sheet.columns:
        target_sheet.column_dimensions[get_column_letter(col[0].column)].width = source_sheet.column_dimensions[
            get_column_letter(col[0].column)].width
    # 保存目标Excel文件为
    if savePath is None:
        return wbDst
    elif savePath == "":
        nfn = m_System.editFilename(wbSrcf, None, "merge-")
        wbDst.save(nfn)
    else:
        wbDst.save(savePath)


def copyWorksheetInto(wbSrcf, wsSrc, wbDstf, savePath="", copyTitle=None):
    """
    复制某Excel到另一个Excel，如果savePath是None，返回wb
    Args:
        wbSrcf:文件1 或者wb对象
        wsSrc:表
        wbDstf: 复制到文件2
        savePath: 保存为
        copyTitle: 复制的标题

    Returns:

    """
    if is_openpyxl_workbook(wbSrcf):
        wbSrc = wbSrcf
    else:
        wbSrc = load_workbook(wbSrcf)
    if is_openpyxl_workbook(wbSrcf):
        wbDst = wbDstf
    else:
        wbDst = load_workbook(wbDstf)
    # 获取源工作表
    source_sheet = getWorksheet(wbSrc, wsSrc)
    # 获取工作表名字，并判断是否要修改
    wsDstName = copyTitle
    if wsDstName is None:
        wsDstName = m_Array.modify_string_if_duplicate(source_sheet.title, getSheetNames(wbDst))
        if wsDstName is None:
            wsDstName = source_sheet.title
    # 新建一张表
    target_sheet = wbDst.create_sheet(title=wsDstName)
    print(f"Copy {wbSrcf} - {source_sheet.title} -> {wbDstf} - {wsDstName}")
    # 复制数据和格式
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]
            target_cell.value = cell.value
            if cell.has_style:
                target_cell._style = cell._style
            if cell.comment:
                target_sheet[cell.coordinate].comment = cell.comment
            # 复制字体
            if source_sheet[cell.coordinate].font:
                target_cell.font = Font(name=source_sheet[cell.coordinate].font.name,
                                        size=source_sheet[cell.coordinate].font.size,
                                        bold=source_sheet[cell.coordinate].font.bold,
                                        italic=source_sheet[cell.coordinate].font.italic,
                                        color=source_sheet[cell.coordinate].font.color)
            # 复制背景色
            if source_sheet[cell.coordinate].fill:
                target_cell.fill = PatternFill(fill_type=source_sheet[cell.coordinate].fill.fill_type,
                                               fgColor=source_sheet[cell.coordinate].fill.start_color.rgb)
            # 复制标注
            if source_sheet[cell.coordinate].comment:
                target_cell.comment = Comment(source_sheet[cell.coordinate].comment.text,
                                              source_sheet[cell.coordinate].comment.author)
    # 复制合并单元格
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))
    # 调整列宽
    for col in source_sheet.columns:
        target_sheet.column_dimensions[get_column_letter(col[0].column)].width = source_sheet.column_dimensions[
            get_column_letter(col[0].column)].width
    # 保存目标Excel文件为3.xlsx
    if savePath is None:
        return wbDst
    elif savePath == "":
        nfn = m_System.editFilename(wbSrcf,None,"merge-")
        wbDst.save(nfn)
    else:
        wbDst.save(savePath)


def copy_cell_format(source_cell, target_cell):
    # 复制源单元格的格式到目标单元格
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def merge_excel_with_format_and_merged_cells(src_folder, output_file, sheet_name, cell_range):
    """
    合并指定文件夹中的Excel文件的某个工作表中的某个范围的数据和格式（包括合并单元格）。

    :param src_folder: 包含Excel文件的文件夹路径
    :param output_file: 保存合并结果的输出文件路径,None则返回wb
    :param sheet_name: 需要读取的工作表名称
    :param cell_range: 需要合并的单元格范围（如'A1:X66'）
    """
    # 创建一个新的工作簿
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = sheet_name
    # 初始化行号，用于在新表中存储数据
    current_row = 1
    # 获取src文件夹中的所有xlsx文件（大小写不敏感）
    files = [f for f in os.listdir(src_folder) if f.lower().endswith('.xlsx')]
    for file in files:
        file_path = os.path.join(src_folder, file)
        # 加载Excel文件
        workbook = load_workbook(file_path, data_only=False)  # 保留格式
        # 获取指定的工作表
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            print(f"警告：文件 {file} 中未找到工作表 {sheet_name}")
            continue
        # 获取要合并的区域，例如 'A1:X66'
        start_cell, end_cell = cell_range.split(':')
        min_col = sheet[start_cell].column
        min_row = sheet[start_cell].row
        max_col = sheet[end_cell].column
        max_row = sheet[end_cell].row
        # 复制合并单元格信息
        for merged_cells in sheet.merged_cells.ranges:
            if merged_cells.min_row >= min_row and merged_cells.max_row <= max_row and \
                    merged_cells.min_col >= min_col and merged_cells.max_col <= max_col:
                new_sheet.merge_cells(
                    start_row=current_row + (merged_cells.min_row - min_row),
                    start_column=merged_cells.min_col,
                    end_row=current_row + (merged_cells.max_row - min_row),
                    end_column=merged_cells.max_col
                )
        # 逐行读取指定区域的内容和格式
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for col_idx, cell in enumerate(row, start=1):
                new_cell = new_sheet.cell(row=current_row, column=col_idx, value=cell.value)
                copy_cell_format(cell, new_cell)  # 复制格式
            current_row += 1
    # 保存合并后的新Excel文件
    if output_file is None:
        return new_workbook
    else:
        new_workbook.save(output_file)
        print(f"合并完成，保存到: {output_file}")


def insert_header_from_excel(target_file_or_wb, header_file, header_range, target_sheets=None, savePath=None):
    """
    从header_file中读取指定区域的表头，并插入到目标Excel文件的所有工作表或指定工作表的顶部。
    表头内容和格式会被保留，并在目标文件原有内容之前插入。

    :param target_file_or_wb: 目标Excel文件路径或wb
    :param header_file: 包含表头的Excel文件路径
    :param header_range: 表头的单元格范围（如'A1:X3'）
    :param target_sheets: 目标工作表的名称列表，默认为None表示插入到所有工作表
    :param savePath: 保存修改后的文件路径，默认为None表示覆盖原文件
    """
    if os.path.isfile(target_file_or_wb):
        # 加载目标Excel文件
        target_wb = load_workbook(target_file_or_wb)
    else:
        target_wb = target_file_or_wb
    # 加载包含表头的Excel文件
    header_wb = load_workbook(header_file)
    header_sheet = header_wb.active  # 默认使用第一个工作表

    # 获取表头范围的起始和结束单元格位置
    min_col, min_row, max_col, max_row = range_boundaries(header_range)
    # 读取表头内容和格式
    header_data = []
    for row in header_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        header_row = []
        for cell in row:
            header_row.append(cell)
        header_data.append(header_row)
    # 记录表头的合并单元格
    header_merged_cells = list(header_sheet.merged_cells.ranges)
    # 确定要插入表头的工作表
    if target_sheets is None:
        # 如果没有指定工作表，则应用到所有工作表
        target_sheets = target_wb.sheetnames
    for sheet_name in target_sheets:
        if sheet_name not in target_wb.sheetnames:
            print(f"警告：目标文件中未找到工作表 {sheet_name}")
            continue
        target_sheet = target_wb[sheet_name]
        # 记录工作表中所有的合并单元格
        merged_cells = list(target_sheet.merged_cells.ranges)
        # 先拆分合并的单元格
        for merged_range in merged_cells:
            target_sheet.unmerge_cells(str(merged_range))
        # 将原有内容下移，插入表头
        target_sheet.insert_rows(1, amount=max_row - min_row + 1)

        # 将表头插入到目标工作表的顶部
        for row_offset, header_row in enumerate(header_data, start=0):
            for col_offset, header_cell in enumerate(header_row, start=0):
                target_cell = target_sheet.cell(row=row_offset + 1, column=min_col + col_offset)
                target_cell.value = header_cell.value
                copy_cell_format(header_cell, target_cell)  # 复制格式
        # 恢复表头的合并单元格
        for merged_range in header_merged_cells:
            min_range_col, min_range_row, max_range_col, max_range_row = range_boundaries(str(merged_range))
            # 调整行号，将表头的合并区域应用到目标表
            new_range = f"{get_column_letter(min_range_col)}{min_range_row}:{get_column_letter(max_range_col)}{max_range_row}"
            target_sheet.merge_cells(new_range)
        # 恢复原有的合并单元格，并调整位置
        for merged_range in merged_cells:
            min_range_col, min_range_row, max_range_col, max_range_row = range_boundaries(str(merged_range))
            # 调整行号，将原有的合并区域向下偏移插入的表头行数
            new_min_row = min_range_row + (max_row - min_row + 1)
            new_max_row = max_range_row + (max_row - min_row + 1)
            new_range = f"{get_column_letter(min_range_col)}{new_min_row}:{get_column_letter(max_range_col)}{new_max_row}"
            # 应用调整后的合并单元格
            target_sheet.merge_cells(new_range)
    # 保存修改后的目标文件
    if savePath is None:
        return target_wb
    else:
        target_wb.save(savePath)
        print(f"表头已插入到目标文件的顶部，保存到: {savePath}")


def remove_blank_rows_or_columns(wb, sheet_name, mode='row'):
    """
    从指定的工作表中移除空白行或空白列。

    :param wb: openpyxl 加载的工作簿对象
    :param sheet_name: 要操作的工作表名称
    :param mode: 'row' 表示移除空白行，'col' 表示移除空白列
    """
    sheet = wb[sheet_name]
    if mode == 'row':
        # 从底部向上删除空白行，防止索引混乱
        for row in range(sheet.max_row, 0, -1):
            if all([cell.value is None for cell in sheet[row]]):
                sheet.delete_rows(row)
    elif mode == 'col':
        # 从右向左删除空白列，防止索引混乱
        for col in range(sheet.max_column, 0, -1):
            if all([sheet.cell(row=row, column=col).value is None for row in range(1, sheet.max_row + 1)]):
                sheet.delete_cols(col)
    print(f"{'空白行' if mode == 'row' else '空白列'} 已移除。")


def remove_blank_cells(wb, sheet_name, cell_range, mode='row'):
    """
    移除指定单元格范围内的空白单元格，按行或列模式处理。
    :param wb: openpyxl 加载的工作簿对象
    :param sheet_name: 要操作的工作表名称
    :param cell_range: 需要处理的单元格范围，例如 'A1:B10'
    :param mode: 'row' 表示按行移除空白单元格并向上移动，'col' 表示按列移除空白单元格并向左移动
    """
    sheet = wb[sheet_name]
    start_cell, end_cell = cell_range.split(':')
    # 获取指定范围的起始和结束位置
    start_col, start_row = sheet[start_cell].column, sheet[start_cell].row
    end_col, end_row = sheet[end_cell].column, sheet[end_cell].row
    if mode == 'row':
        # 按行模式移除空白单元格
        for col in range(start_col, end_col + 1):
            non_empty_values = [sheet.cell(row=row, column=col).value for row in range(start_row, end_row + 1) if
                                sheet.cell(row=row, column=col).value is not None]
            for i, value in enumerate(non_empty_values):
                sheet.cell(row=start_row + i, column=col).value = value
            for row in range(start_row + len(non_empty_values), end_row + 1):
                sheet.cell(row=row, column=col).value = None
    elif mode == 'col':
        # 按列模式移除空白单元格
        for row in range(start_row, end_row + 1):
            non_empty_values = [sheet.cell(row=row, column=col).value for col in range(start_col, end_col + 1) if
                                sheet.cell(row=row, column=col).value is not None]
            for i, value in enumerate(non_empty_values):
                sheet.cell(row=row, column=start_col + i).value = value
            for col in range(start_col + len(non_empty_values), end_col + 1):
                sheet.cell(row=row, column=col).value = None
    print(f"{'行' if mode == 'row' else '列'}模式下的空白单元格已移除。")


if __name__ == '__main__':
    print("Hello World")

